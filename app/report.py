"""
Module: report
Contains the ExcelReport class for exporting the log and summary to an Excel file.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from .data_manager import load_data
from datetime import datetime
from .constants import MONTH_NAMES, DEFAULT_AMOUNT

class ExcelReport:
    """
    A class to generate an Excel report from the sobriety log.
    """
    def __init__(self, data: dict):
        self.data = data

    def generate(self, filename: str) -> None:
        """
        Generates and saves an Excel report with log entries and a summary.

        Parameters:
            filename (str): The name of the Excel file to create.
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sobriety Report"
            ws.append(["Date", "Status", "Mood", "Type of Alcohol", "Amount Consumed", "Notes", "Amount Spent"])

            daily_log = self.data.get("daily_log", {})
            for date_str in sorted(daily_log.keys()):
                entry = daily_log[date_str]
                status_text = "Sober" if entry.get("sober", False) else "Drinking"
                if entry.get("sober", False):
                    amount_spent_str = ""
                else:
                    try:
                        amount_spent = float(entry.get("amount_spent", DEFAULT_AMOUNT))
                        if amount_spent == 0:
                            amount_spent = DEFAULT_AMOUNT
                    except Exception:
                        amount_spent = DEFAULT_AMOUNT
                    amount_spent_str = f"{amount_spent:.2f} $"
                ws.append([
                    date_str,
                    status_text,
                    entry.get("mood", ""),
                    entry.get("alcohol_type", ""),
                    entry.get("alcohol_amount", ""),
                    entry.get("notes", ""),
                    amount_spent_str
                ])

            for column_cells in ws.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
                ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

            # Format status column
            light_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            light_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
                for cell in row:
                    if cell.value == "Sober":
                        cell.fill = light_green
                    elif cell.value == "Drinking":
                        cell.fill = light_red

            # Summary calculations
            total_days = len(daily_log)
            sober_days = sum(1 for record in daily_log.values() if record.get("sober", False))
            drinking_days = total_days - sober_days
            percent_sober = (sober_days / total_days * 100) if total_days > 0 else 0

            sorted_dates = sorted(daily_log.keys())
            longest_streak = 0
            current_streak = 0
            previous_date = None
            for date_str in sorted_dates:
                try:
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                except Exception:
                    continue
                if daily_log[date_str].get("sober", False):
                    if previous_date and (date_obj - previous_date).days == 1:
                        current_streak += 1
                    else:
                        current_streak = 1
                    longest_streak = max(longest_streak, current_streak)
                else:
                    current_streak = 0
                previous_date = date_obj

            current_seq = 0
            for date_str in reversed(sorted_dates):
                if daily_log[date_str].get("sober", False):
                    current_seq += 1
                else:
                    break

            overall_spending = sum(
                float(record.get("amount_spent", DEFAULT_AMOUNT))
                if (not record.get("sober", True)) and float(record.get("amount_spent", DEFAULT_AMOUNT)) > 0
                else DEFAULT_AMOUNT
                for record in daily_log.values() if not record.get("sober", False)
            )

            summary_start_col = 8  # Column H
            summary_start_row = 2

            ws.merge_cells(start_row=summary_start_row, start_column=summary_start_col,
                           end_row=summary_start_row, end_column=summary_start_col+1)
            header_cell = ws.cell(row=summary_start_row, column=summary_start_col)
            header_cell.value = "Summary"
            header_cell.font = Font(bold=True, size=14)
            header_cell.alignment = Alignment(horizontal="center", wrap_text=True)
            header_cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

            label_font = Font(bold=True)
            data_alignment = Alignment(horizontal="left", wrap_text=True)

            ws.cell(row=summary_start_row+1, column=summary_start_col, value="Total days in log:").font = label_font
            ws.cell(row=summary_start_row+1, column=summary_start_col).alignment = data_alignment
            ws.cell(row=summary_start_row+1, column=summary_start_col+1, value=total_days)

            ws.cell(row=summary_start_row+2, column=summary_start_col, value="Sober days:").font = label_font
            ws.cell(row=summary_start_row+2, column=summary_start_col).alignment = data_alignment
            ws.cell(row=summary_start_row+2, column=summary_start_col+1, value=sober_days)

            ws.cell(row=summary_start_row+3, column=summary_start_col, value="Drinking days:").font = label_font
            ws.cell(row=summary_start_row+3, column=summary_start_col).alignment = data_alignment
            ws.cell(row=summary_start_row+3, column=summary_start_col+1, value=drinking_days)

            ws.cell(row=summary_start_row+4, column=summary_start_col, value="Percentage sober:").font = label_font
            ws.cell(row=summary_start_row+4, column=summary_start_col).alignment = data_alignment
            ws.cell(row=summary_start_row+4, column=summary_start_col+1, value=f"{percent_sober:.2f}%")

            ws.cell(row=summary_start_row+5, column=summary_start_col, value="Longest sober streak:").font = label_font
            ws.cell(row=summary_start_row+5, column=summary_start_col).alignment = data_alignment
            ws.cell(row=summary_start_row+5, column=summary_start_col+1, value=f"{longest_streak} days")

            ws.cell(row=summary_start_row+6, column=summary_start_col, value="Current sober streak:").font = label_font
            ws.cell(row=summary_start_row+6, column=summary_start_col).alignment = data_alignment
            ws.cell(row=summary_start_row+6, column=summary_start_col+1, value=f"{current_seq} days")

            ws.cell(row=summary_start_row+7, column=summary_start_col, value="Total amount spent (overall):").font = label_font
            ws.cell(row=summary_start_row+7, column=summary_start_col).alignment = data_alignment
            ws.cell(row=summary_start_row+7, column=summary_start_col+1, value=f"{overall_spending:.2f} $")

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            for row in range(summary_start_row, summary_start_row+8):
                for col in range(summary_start_col, summary_start_col+2):
                    ws.cell(row=row, column=col).border = thin_border

            ws.column_dimensions['H'].width = 35
            ws.column_dimensions['I'].width = 25

            wb.save(filename)
            print(f"Report saved to file: {filename}")
        except Exception as e:
            print(f"Error generating report: {e}")
